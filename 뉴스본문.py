import requests
import re
from openpyxl import load_workbook

# API endpoint for news content retrieval
url = "https://tools.kinds.or.kr/search/news"

def preprocess_title(title):
    # [단독], (종합) 등 제거 및 특수문자 및 한자 처리
    title = re.sub(r'\[.*?\]', '', title)  # [ ] 내용 제거
    title = re.sub(r'\(.*?\)', '', title)  # ( ) 내용 제거
    title = re.sub(r'[\u4e00-\u9fff]', '', title)  # 한자 제거
    title = re.sub(r'[\"\'‘’“”…···]', '', title)  # 특수문자 제거 ("'… 포함)
    title = re.sub(r'[\,\;\:\!\?]', '', title)  # 쉼표 등 제거

    title = title.strip()
    return title

def get_news_content_by_title(title):
    # Replace with your actual access key
    access_key = "e8cf85e5-a3ce-467e-8187-96183dbe6bf7"

    # Request payload
    payload = {
        "access_key": access_key,
        "argument": {
            "query": title,  # Search for the news by title
            "published_at": {
                "from": "2000-10-31",  # Start date
                "until": "2024-11-28"  # End date
            },
            "fields": ["content", "title", "provider"],  # Specify fields you need
            "return_size": 1  # Limit to the first result
        }
    }

    try:
        # Make the POST request
        response = requests.post(url, json=payload)
        response.raise_for_status()

        # Parse response
        data = response.json()
        if data["result"] == 0 and "documents" in data["return_object"]:
            documents = data["return_object"]["documents"]
            if documents:
                content = documents[0].get("content", "Content not available")
                return content
            else:
                return "No matching news article found."
        else:
            return f"Error in response: {data}"
    except Exception as e:
        return f"An error occurred: {str(e)}"

def retry_no_matching_articles():
    # 기존 데이터 로드
    output_file = "C:\\Users\\강예린\\본문.xlsx"
    workbook = load_workbook(output_file)
    sheet = workbook.active

    # 행 데이터 읽기
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]

    for row_idx, row in enumerate(data, start=2):  # start=2 to account for header
        title, content = row
        if content == "No matching news article found." and "···" in title:
            preprocessed_title = preprocess_title(title)
            new_content = get_news_content_by_title(preprocessed_title)
            sheet.cell(row=row_idx, column=2, value=new_content)  # Update content column
            workbook.save(output_file)
            print(f"Updated: {preprocess_title(title)} {new_content}")

    print(f"No matching articles 처리 완료")

# 실행
retry_no_matching_articles()
